from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from railwaystation import BackwardConstructionAlgorithm, DataProvider, StandardCaseConverter, Terminal


BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "RailwayStation_V2_py" / "Data"
RUNNER_OUTPUT_DIR = BASE_DIR / "RailwayStation_V2.Runner"

SAMPLE_NAMES = [
    "2025-09-04-noon",
    "2025-09-08-noon",
    "2025-09-08-afternoon",
    "2025-09-09-noon",
    "2025-11-04-noon",
    "2025-11-04-afternoon",
    "2025-11-05-noon",
    "2025-11-05-afternoon",
    "2025-11-06-noon",
    "2025-11-06-afternoon",
    "2025-11-07-noon",
    "2025-11-10-noon",
    "2025-11-10-afternoon",
    "2025-11-11-noon",
    "2025-12-08-noon",
    "2025-12-08-afternoon",
    "2025-12-09-noon",
    "2025-12-09-afternoon",
]


def prepare_case_for_solver(sample_name: str) -> Path:
    case_path = DATA_DIR / f"{sample_name}.xlsx"
    working_case_path = Path(StandardCaseConverter.convert_case(str(case_path)))
    workbook = load_workbook(working_case_path)
    has_end_generated = "End_generated" in workbook.sheetnames
    workbook.close()
    if not has_end_generated:
        Terminal.generate_end_sheet(str(case_path))
    return working_case_path


def build_python_operation_signatures(sample_name: str) -> list[dict[str, object]]:
    distance_matrix, track_line_capacity = DataProvider.get_map_info(str(DATA_DIR / "map.xlsx"))
    track_lines = DataProvider.init_track_lines(track_line_capacity)
    cars = DataProvider.init_cars(track_lines, str(prepare_case_for_solver(sample_name)))
    operations = BackwardConstructionAlgorithm(track_lines, cars, distance_matrix).run()
    return [
        {
            "line": op.line_name,
            "action": op.action.value,
            "cars": [car.no for car in op.move_cars],
        }
        for op in operations
    ]


def iter_current_source_artifacts() -> Iterable[tuple[str, Path]]:
    for sample_name in SAMPLE_NAMES:
        result_path = RUNNER_OUTPUT_DIR / f"output_{sample_name}_std" / "direct_solver_only.json"
        if result_path.exists():
            yield sample_name, result_path


def build_csharp_operation_signatures(result_path: Path) -> list[dict[str, object]]:
    operations = json.loads(result_path.read_text())["Operations"]
    return [
        {
            "line": op["LineName"],
            "action": op["Action"],
            "cars": [car["No"] for car in op["MoveCars"]],
        }
        for op in operations
    ]


def compare_sample(sample_name: str, result_path: Path) -> None:
    py_ops = build_python_operation_signatures(sample_name)
    cs_ops = build_csharp_operation_signatures(result_path)
    assert len(py_ops) == len(cs_ops), f"{sample_name}: operation count mismatch {len(py_ops)} != {len(cs_ops)}"
    for index, (py_op, cs_op) in enumerate(zip(py_ops, cs_ops), start=1):
        assert py_op == cs_op, f"{sample_name}: operation #{index} mismatch\npy={py_op}\ncs={cs_op}"


def test_all_bundled_samples_match_csharp_artifacts() -> None:
    available_artifacts = list(iter_current_source_artifacts())
    assert available_artifacts, "No current-source direct_solver_only artifacts found under RailwayStation_V2.Runner/output_*_std"
    for sample_name, result_path in available_artifacts:
        compare_sample(sample_name, result_path)
