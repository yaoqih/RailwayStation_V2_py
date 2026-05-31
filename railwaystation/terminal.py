from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

from .core import Car, CarGroup, TrackLine
from .data_provider import DataProvider
from .io import TerminalContext
from .standard_converter import StandardCaseConverter
from .terminal_strategies import SolverBlindSpotAvoidanceTerminalStrategy
from .terminal_strategy_runner import TerminalStrategyRunner


@dataclass
class PlacementGroup:
    source_line: TrackLine
    cars: list[Car]
    start_index: int
    blocker_count: int
    distance: int
    is_same_line: bool
    chunk_index: int
    source_order: int

    SourceLine = property(lambda self: self.source_line, lambda self, value: setattr(self, "source_line", value))
    Cars = property(lambda self: self.cars, lambda self, value: setattr(self, "cars", value))
    StartIndex = property(lambda self: self.start_index, lambda self, value: setattr(self, "start_index", value))
    BlockerCount = property(lambda self: self.blocker_count, lambda self, value: setattr(self, "blocker_count", value))
    Distance = property(lambda self: self.distance, lambda self, value: setattr(self, "distance", value))
    IsSameLine = property(lambda self: self.is_same_line, lambda self, value: setattr(self, "is_same_line", value))
    ChunkIndex = property(lambda self: self.chunk_index, lambda self, value: setattr(self, "chunk_index", value))
    SourceOrder = property(lambda self: self.source_order, lambda self, value: setattr(self, "source_order", value))


@dataclass
class GroupPlacementPlan:
    group: PlacementGroup
    has_forced_car: bool
    can_use_exact_block: bool = False
    exact_start: int = -1
    exact_end: int = -1

    Group = property(lambda self: self.group, lambda self, value: setattr(self, "group", value))
    HasForcedCar = property(lambda self: self.has_forced_car, lambda self, value: setattr(self, "has_forced_car", value))
    CanUseExactBlock = property(lambda self: self.can_use_exact_block, lambda self, value: setattr(self, "can_use_exact_block", value))
    ExactStart = property(lambda self: self.exact_start, lambda self, value: setattr(self, "exact_start", value))
    ExactEnd = property(lambda self: self.exact_end, lambda self, value: setattr(self, "exact_end", value))


class Terminal:
    max_continuous_chunk_size = 5
    repair_outer_start_position = 1
    repair_outer_capacity = 4
    repair_inner_start_position = 5
    repair_inner_capacity = 5

    @staticmethod
    def _display_width(value) -> int:
        text = "" if value is None else str(value)
        width = 0
        for ch in text:
            width += 2 if unicodedata.east_asian_width(ch) in {"F", "W"} else 1
        return width

    @staticmethod
    def _used_row_numbers(sheet) -> list[int]:
        result: list[int] = []
        for row in sheet.iter_rows():
            values = tuple(cell.value for cell in row)
            if DataProvider.is_effectively_empty_row(values):
                continue
            result.append(row[0].row)
        return result

    @staticmethod
    def _last_used_column(sheet) -> int:
        last_col = 0
        for row in sheet.iter_rows():
            values = tuple(cell.value for cell in row)
            if DataProvider.is_effectively_empty_row(values):
                continue
            for cell in row:
                if cell.value is None or str(cell.value).strip() == "":
                    continue
                if cell.column > last_col:
                    last_col = cell.column
        return last_col

    @staticmethod
    def generate_end_sheet(file_path: str) -> None:
        standardized_path = Terminal.prepare_standardized_case(file_path)
        Terminal.ensure_start_with_end_sheet(str(standardized_path))
        context = TerminalContext.build_terminal_context(str(standardized_path))
        distance_matrix = context.distance_matrix
        track_line_capacity = context.track_line_capacity
        best_result = TerminalStrategyRunner.find_best_solve(
            str(standardized_path),
            distance_matrix,
            track_line_capacity,
        )
        context = best_result.context
        if True:
            Terminal.testCar(context)
            Terminal.copy_source_file(str(standardized_path))
            Terminal.output_file(str(standardized_path), context)

    @staticmethod
    def prepare_standardized_case(file_path: str) -> Path:
        return StandardCaseConverter.convert_case(file_path)

    @staticmethod
    def ensure_start_with_end_sheet(file_path: str) -> None:
        workbook = load_workbook(file_path)
        need_rebuild = True
        if "Start_with_end" in workbook.sheetnames:
            sheet = workbook["Start_with_end"]
            need_rebuild = not Terminal._has_valid_header(sheet)
        workbook.close()
        if need_rebuild:
            Terminal.add_end_position_to_start_sheet(file_path)

    @staticmethod
    def _has_valid_header(sheet) -> bool:
        headers = []
        for cell in sheet[1]:
            value = str(cell.value or "").strip()
            if value:
                headers.append(value)
        required = {"股道", "序号", "车型", "车号", "末尾位置", "强制对位"}
        return required.issubset(set(headers))

    @staticmethod
    def _find_header_row(sheet) -> int:
        for row in sheet.iter_rows():
            headers = [str(cell.value or "").strip() for cell in row if str(cell.value or "").strip()]
            if {"股道", "序号", "车型", "车号"}.issubset(set(headers)):
                return row[0].row
        raise RuntimeError(f"工作表 {sheet.title} 中未找到表头行")

    @staticmethod
    def testCar(context: TerminalContext) -> None:
        for car in context.cars.values():
            print(
                f"{car.no}原来在{car.origin_line_name},台号为{car.origin_line_position},"
                f"被分配到{car.target_line_name}，台号为{car.target_line_position}"
            )

    @staticmethod
    def output_file(filepath: str, context: TerminalContext) -> bool:
        flag = True
        workbook = load_workbook(filepath)
        sheet = DataProvider.get_sheet(workbook, "End_generated")
        used_rows = [row for row in Terminal._used_row_numbers(sheet) if row >= 2]
        for row in used_rows:
            sheet.cell(row, 3).value = None
            sheet.cell(row, 4).value = None
        for car in context.cars.values():
            if not car.target_line_name or car.target_line_position < 0:
                continue
            matched_row = None
            for row in used_rows:
                if str(sheet.cell(row, 1).value or "").strip() == car.target_line_name and str(sheet.cell(row, 2).value or "").strip() == str(car.target_line_position):
                    matched_row = row
                    break
            if matched_row is not None:
                sheet.cell(matched_row, 3).value = car.type
                sheet.cell(matched_row, 4).value = car.no
            else:
                flag = False
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = 12
        for row_idx, row in enumerate(sheet.iter_rows(), start=1):
            sheet.row_dimensions[row_idx].height = 20
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        workbook.save(filepath)
        return flag

    @staticmethod
    def copy_source_file(file_path: str) -> None:
        workbook = load_workbook(file_path)
        source_sheet = DataProvider.get_sheet(workbook, "End")
        if "End_generated" in workbook.sheetnames:
            del workbook["End_generated"]
        end_sheet = workbook.create_sheet("End_generated")
        last_col = Terminal._last_used_column(source_sheet)
        new_row = 1
        for row_number in Terminal._used_row_numbers(source_sheet):
            for col in range(1, last_col + 1):
                end_sheet.cell(new_row, col).value = source_sheet.cell(row_number, col).value
            new_row += 1
        workbook.save(file_path)

    @staticmethod
    def add_end_position_to_start_sheet(filepath: str) -> None:
        workbook = load_workbook(filepath)
        start_sheet = DataProvider.get_sheet(workbook, "Start")
        # V3 的 AddEndPositionToStartSheet 依赖终点模板语义，而不是上一次运行后可能被污染的 End_generated。
        # 这里优先使用稳定的 End 工作表，避免把半成品回写成新的 Start_with_end。
        end_sheet = DataProvider.get_sheet(workbook, "End")
        start_header_row = Terminal._find_header_row(start_sheet)
        end_header_row = Terminal._find_header_row(end_sheet)
        end_force_column = 0
        end_sequence_column = 2
        end_line_column = 1
        end_car_no_column = 4
        end_last_col = Terminal._last_used_column(end_sheet)
        for col in range(1, end_last_col + 1):
            header = str(end_sheet.cell(end_header_row, col).value or "").strip()
            if header == "强制对位":
                end_force_column = col
            elif header == "序号":
                end_sequence_column = col
            elif header == "股道":
                end_line_column = col
            elif header == "车号":
                end_car_no_column = col
        end_position_dict: dict[str, str] = {}
        end_target_position_dict: dict[str, str] = {}
        for row in [row for row in Terminal._used_row_numbers(end_sheet) if row > end_header_row]:
            line_name = str(end_sheet.cell(row, end_line_column).value or "").strip()
            position = str(end_sheet.cell(row, end_sequence_column).value or "").strip()
            car_no = str(end_sheet.cell(row, end_car_no_column).value or "").strip()
            if not car_no:
                continue
            end_position_dict[car_no] = Terminal.map_end_line_name_by_position(line_name, position)
            end_target_position_dict[car_no] = position
        if "Start_with_end" in workbook.sheetnames:
            del workbook["Start_with_end"]
        new_sheet = workbook.create_sheet("Start_with_end")
        last_col = Terminal._last_used_column(start_sheet)
        for col in range(1, last_col + 1):
            new_sheet.cell(1, col).value = start_sheet.cell(start_header_row, col).value
        new_col = last_col + 1
        new_sheet.cell(1, new_col).value = "末尾位置"
        new_sheet.cell(1, new_col + 1).value = "强制对位"
        start_car_no_column = 4
        for col in range(1, last_col + 1):
            header = str(start_sheet.cell(start_header_row, col).value or "").strip()
            if header == "车号":
                start_car_no_column = col
                break
        new_row_index = 2
        for row in [row for row in Terminal._used_row_numbers(start_sheet) if row > start_header_row]:
            for col in range(1, last_col + 1):
                new_sheet.cell(new_row_index, col).value = start_sheet.cell(row, col).value
            car_no = str(start_sheet.cell(row, start_car_no_column).value or "").strip()
            new_sheet.cell(new_row_index, new_col).value = end_position_dict.get(car_no, "未找到")
            new_sheet.cell(new_row_index, new_col + 1).value = end_target_position_dict.get(car_no, "0")
            new_row_index += 1
        for column_cells in new_sheet.columns:
            column_letter = column_cells[0].column_letter
            width = max((Terminal._display_width(cell.value) for cell in column_cells), default=0) + 2
            if width < 15:
                width = 15
            if width > 50:
                width = 50
            new_sheet.column_dimensions[column_letter].width = width
        workbook.save(filepath)

    @staticmethod
    def build_global_dict(track_line_dict: dict[str, TrackLine]) -> dict[str, dict[str, list[CarGroup]]]:
        global_dict: dict[str, dict[str, list[CarGroup]]] = {}
        for track_name, track_line in track_line_dict.items():
            terminal_dict = Terminal.get_target_line_segment(track_line)
            for target_line_name, groups in terminal_dict.items():
                global_dict.setdefault(target_line_name, {})
                global_dict[target_line_name].setdefault(track_name, [])
                global_dict[target_line_name][track_name].extend(groups)
        return global_dict

    @staticmethod
    def get_target_line_segment(track_line: TrackLine) -> dict[str, list[CarGroup]]:
        terminal_dict: dict[str, list[CarGroup]] = {}
        if not track_line or not track_line.current_list:
            return terminal_dict
        cars = track_line.current_list
        groups: list[CarGroup] = []
        i = 0
        while i < len(cars):
            current_car = cars[i]
            group = CarGroup()
            group.cars.append(current_car)
            j = i + 1
            while j >= 0 and j < len(cars) and cars[j].target_line_name == current_car.target_line_name:
                group.cars.append(cars[j])
                j += 1
            groups.append(group)
            i = j
        for group in groups:
            target_line_name = group.top_car.target_line_name
            terminal_dict.setdefault(target_line_name, []).append(group)
        return terminal_dict

    @staticmethod
    def assign_position_direct(global_dict: dict[str, dict[str, list[CarGroup]]], context: TerminalContext) -> bool:
        for target_line_name, source_groups in global_dict.items():
            target_track = context.track_lines[target_line_name]
            target_track.target_list.clear()
            placement_groups = Terminal.build_placement_groups(source_groups, context, target_line_name)
            max_position = context.start_position_dict[target_line_name]
            all_cars = [car for group in placement_groups for car in group.cars]
            if not all_cars:
                continue
            if len(all_cars) > max_position + 1:
                raise RuntimeError("剩余台位不够分")
            if any(car.is_force_target_position for car in all_cars):
                success, assigned_position_dict = Terminal.try_build_force_aware_assigned_position_dict(placement_groups, max_position)
                if not success:
                    assigned_position_dict = Terminal.build_assigned_position_dict(all_cars, max_position)
            else:
                assigned_position_dict = Terminal.build_assigned_position_dict(all_cars, max_position)
            for car in all_cars:
                car.target_line = target_track
                car.target_line_name = target_track.name
                car.target_line_position = assigned_position_dict[car]
            for car in sorted(all_cars, key=lambda c: (c.target_line_position, c.origin_line_position)):
                target_track.unshift_target(car)
        return True

    @staticmethod
    def build_assigned_position_dict(cars: list[Car], max_position: int) -> dict[Car, int]:
        result: dict[Car, int] = {}
        forced_position_dict: dict[int, Car] = {}
        for car in cars:
            if not car.is_force_target_position:
                continue
            fixed_position = car.fixed_target_line_position
            if fixed_position < 0 or fixed_position > max_position:
                raise RuntimeError(f"车辆{car.no}的强制对位台号{fixed_position}超出允许范围[0,{max_position}]")
            if fixed_position in forced_position_dict:
                existed_car = forced_position_dict[fixed_position]
                raise RuntimeError(f"目标线存在重复强制对位台号：{fixed_position}，车辆{existed_car.no}与{car.no}冲突")
            forced_position_dict[fixed_position] = car
            result[car] = fixed_position
        remain_positions = Terminal.find_compact_remain_positions(max_position, set(forced_position_dict.keys()), len(cars) - len(forced_position_dict))
        remain_index = 0
        for car in cars:
            if car.is_force_target_position:
                continue
            result[car] = remain_positions[remain_index]
            remain_index += 1
        return result

    @staticmethod
    def try_build_force_aware_assigned_position_dict(placement_groups: list[PlacementGroup], max_position: int) -> tuple[bool, dict[Car, int]]:
        result: dict[Car, int] = {}
        if not placement_groups:
            return True, result
        plans = Terminal.build_group_placement_plans(placement_groups, max_position)
        if any(plan.has_forced_car and not plan.can_use_exact_block for plan in plans):
            return False, {}
        occupied_positions = [False] * (max_position + 1)
        exact_plans = sorted([plan for plan in plans if plan.can_use_exact_block], key=lambda p: (p.exact_start, p.exact_end))
        for plan in exact_plans:
            for position in range(plan.exact_start, plan.exact_end + 1):
                if occupied_positions[position]:
                    return False, {}
            for i, car in enumerate(plan.group.cars):
                position = plan.exact_start + i
                if car.is_force_target_position and car.fixed_target_line_position != position:
                    return False, {}
                occupied_positions[position] = True
                result[car] = position
        unanchored_plans = [plan for plan in plans if not plan.can_use_exact_block]
        unanchored_plans.sort(key=lambda p: (p.group.is_same_line, -len(p.group.cars), Terminal.ease_key(p.group)))
        for plan in unanchored_plans:
            if not Terminal.try_assign_unanchored_group(plan, occupied_positions, max_position, result):
                return False, {}
        return True, result

    @staticmethod
    def build_group_placement_plans(placement_groups: list[PlacementGroup], max_position: int) -> list[GroupPlacementPlan]:
        result: list[GroupPlacementPlan] = []
        for placement_group in placement_groups:
            forced_cars = [(car, index) for index, car in enumerate(placement_group.cars) if car.is_force_target_position]
            plan = GroupPlacementPlan(group=placement_group, has_forced_car=bool(forced_cars))
            if not forced_cars:
                result.append(plan)
                continue
            start_candidates: list[int] = []
            for car, index in forced_cars:
                candidate = car.fixed_target_line_position - index
                if candidate not in start_candidates:
                    start_candidates.append(candidate)
            if len(start_candidates) == 1:
                exact_start = start_candidates[0]
                exact_end = exact_start + len(placement_group.cars) - 1
                if exact_start >= 0 and exact_end <= max_position:
                    plan.can_use_exact_block = True
                    plan.exact_start = exact_start
                    plan.exact_end = exact_end
            result.append(plan)
        return result

    @staticmethod
    def try_assign_unanchored_group(plan: GroupPlacementPlan, occupied_positions: list[bool], max_position: int, result: dict[Car, int]) -> bool:
        group_length = len(plan.group.cars)
        free_intervals = Terminal.get_free_intervals(occupied_positions, max_position)
        best_block = None
        best_adjacency = -1
        best_end = -1
        best_start = -1
        for start, end in free_intervals:
            interval_length = end - start + 1
            if interval_length < group_length:
                continue
            candidates = {(start, start + group_length - 1), (end - group_length + 1, end)}
            for candidate_start, candidate_end in candidates:
                adjacency = 0
                if candidate_start > 0 and occupied_positions[candidate_start - 1]:
                    adjacency += 1
                if candidate_end < max_position and occupied_positions[candidate_end + 1]:
                    adjacency += 1
                if adjacency > best_adjacency or (adjacency == best_adjacency and candidate_end > best_end) or (
                    adjacency == best_adjacency and candidate_end == best_end and candidate_start > best_start
                ):
                    best_adjacency = adjacency
                    best_end = candidate_end
                    best_start = candidate_start
                    best_block = (candidate_start, candidate_end)
        if best_block is None:
            return False
        for i, car in enumerate(plan.group.cars):
            position = best_block[0] + i
            occupied_positions[position] = True
            result[car] = position
        return True

    @staticmethod
    def get_free_intervals(occupied_positions: list[bool], max_position: int) -> list[tuple[int, int]]:
        result: list[tuple[int, int]] = []
        start = None
        for position in range(0, max_position + 1):
            if not occupied_positions[position]:
                if start is None:
                    start = position
            elif start is not None:
                result.append((start, position - 1))
                start = None
        if start is not None:
            result.append((start, max_position))
        return result

    @staticmethod
    def find_compact_remain_positions(max_position: int, forced_positions: set[int], remain_count: int) -> list[int]:
        result: list[int] = []
        if remain_count <= 0:
            return result
        forced_min = min(forced_positions) if forced_positions else -1
        forced_max = max(forced_positions) if forced_positions else -1
        best_start = -1
        best_end = -1
        best_span = 10**18
        for start in range(0, max_position + 1):
            if forced_positions and start > forced_min:
                break
            end_start = max(start, forced_max) if forced_positions else start
            for end in range(end_start, max_position + 1):
                free_count = sum(1 for pos in range(start, end + 1) if pos not in forced_positions)
                if free_count < remain_count:
                    continue
                span = end - start + 1
                if span < best_span or (span == best_span and end > best_end) or (span == best_span and end == best_end and start > best_start):
                    best_span = span
                    best_start = start
                    best_end = end
        if best_start < 0:
            raise RuntimeError("剩余台位不够分")
        for pos in range(best_start, best_end + 1):
            if pos not in forced_positions:
                result.append(pos)
        if len(result) < remain_count:
            raise RuntimeError("剩余台位不够分")
        return result[:remain_count]

    @staticmethod
    def resolve_group_segment(cars: list[Car], line_max_position: int) -> tuple[int, int]:
        min_candidates = [car.target_min_position for car in cars if car.target_min_position > 0]
        max_candidates = [car.target_max_position for car in cars if car.target_max_position > 0]
        min_position = max(min_candidates) if min_candidates else 1
        max_position = min(min(max_candidates), line_max_position) if max_candidates else line_max_position
        return min_position, max_position

    @staticmethod
    def get_source_priority(source_line_name: str, use_normalized: bool = False) -> int:
        priority_key = Terminal.normalize_track_name_for_distance(source_line_name) if use_normalized else source_line_name
        if not priority_key:
            return 9
        if priority_key in {"洗罐", "喷漆"}:
            return 0
        if priority_key == "卸轮线":
            return 1
        if priority_key.startswith("修"):
            return 1
        if priority_key == "机走":
            return 2
        if priority_key in {"机库线", "调梁"}:
            return 3
        if priority_key == "老预修":
            return 5
        if priority_key.startswith("存"):
            return 6
        return 4

    @staticmethod
    def build_placement_groups(source_groups: dict[str, list[CarGroup]], context: TerminalContext, target_line_name: str) -> list[PlacementGroup]:
        raw_groups = []
        for source in source_groups.values():
            for index, group in enumerate(source):
                raw_groups.append(Terminal.create_placement_group(group, context, target_line_name, index))
        if not raw_groups:
            return []
        external_groups = [group for group in raw_groups if not group.is_same_line]
        same_line_groups = [group for group in raw_groups if group.is_same_line]
        if not external_groups:
            return sorted(same_line_groups, key=lambda g: (g.source_order, g.start_index))
        can_interleave_chunks = len({g.source_line.name for g in external_groups}) > 1 or bool(same_line_groups)
        if can_interleave_chunks:
            external_groups = [item for group in external_groups for item in Terminal.split_placement_group(group)]
        shallow_to_deep: list[PlacementGroup] = []
        source_buckets = []
        for _, grouped in sorted(
            ((key, list(value)) for key, value in Terminal.group_by_source(external_groups).items()),
            key=lambda item: Terminal.ease_key(sorted(item[1], key=lambda g: (g.chunk_index, g.start_index, g.source_order))[0]),
        ):
            source_buckets.append(sorted(grouped, key=lambda g: (g.chunk_index, g.start_index, g.source_order)))
        has_group = True
        while has_group:
            has_group = False
            for bucket in source_buckets:
                if not bucket:
                    continue
                shallow_to_deep.append(bucket.pop(0))
                has_group = True
        same_line_groups = sorted(same_line_groups, key=lambda g: (g.source_order, g.start_index))
        shallow_to_deep.extend(same_line_groups)
        return shallow_to_deep

    @staticmethod
    def create_placement_group(group: CarGroup, context: TerminalContext, target_line_name: str, source_order: int) -> PlacementGroup:
        top_car = group.top_car
        return PlacementGroup(
            source_line=group.current_line,
            cars=list(group.cars),
            start_index=top_car.origin_line_position,
            blocker_count=top_car.current_depth,
            distance=Terminal.get_distance(context.distance_matrix, group.current_line.name, target_line_name),
            is_same_line=group.current_line.name == target_line_name,
            chunk_index=0,
            source_order=source_order,
        )

    @staticmethod
    def split_placement_group(group: PlacementGroup) -> list[PlacementGroup]:
        if len(group.cars) <= Terminal.max_continuous_chunk_size or any(car.is_force_target_position for car in group.cars):
            return [group]
        result: list[PlacementGroup] = []
        chunk_index = 0
        for i in range(0, len(group.cars), Terminal.max_continuous_chunk_size):
            cars = group.cars[i : i + Terminal.max_continuous_chunk_size]
            result.append(
                PlacementGroup(
                    source_line=group.source_line,
                    cars=cars,
                    start_index=cars[0].origin_line_position,
                    blocker_count=group.blocker_count + i,
                    distance=group.distance,
                    is_same_line=group.is_same_line,
                    chunk_index=chunk_index,
                    source_order=group.source_order,
                )
            )
            chunk_index += 1
        return result

    @staticmethod
    def get_distance(distance_matrix: dict[str, dict[str, int]], source_line_name: str, target_line_name: str) -> int:
        source_key = Terminal.normalize_track_name_for_distance(source_line_name)
        target_key = Terminal.normalize_track_name_for_distance(target_line_name)
        return distance_matrix.get(source_key, {}).get(target_key, 2**31 - 1)

    @staticmethod
    def normalize_track_name_for_distance(name: str) -> str:
        mapping = {
            "老预修": "老预修",
            "喷漆库": "喷漆",
            "喷漆库外": "喷漆",
            "喷漆库内": "喷漆",
            "机库线": "机库线",
            "机走预修": "机走",
            "调梁库": "调梁",
            "修1库外": "修1",
            "修1库内": "修1",
            "修2库外": "修2",
            "修2库内": "修2",
            "卸轮线": "卸轮线",
            "存1线": "存1线",
            "存2线": "存2线",
            "存3线": "存3线",
            "存4线": "存4线",
            "存5线北": "存5线",
            "存5线南": "存5线",
            "抛丸线": "抛丸线",
            "洗罐线": "洗罐",
            "洗罐线外": "洗罐",
            "洗罐线内": "洗罐",
            "调梁库外": "调梁",
            "调梁库内": "调梁",
            "修3库外": "修3",
            "修3库内": "修3",
            "修4库外": "修4",
            "修4库内": "修4",
            "洗罐库外": "洗罐",
            "洗罐库内": "洗罐",
        }
        return mapping.get(name, name)

    @staticmethod
    def map_end_line_name_by_position(line_name: str, position_text: str) -> str:
        line_name = line_name.strip()
        try:
            position = int(position_text.strip())
        except Exception:
            return line_name
        mapping = {
            "老预修": "预修线",
            "机库线": "机库线",
            "卸轮线": "卸轮线",
            "喷漆": "油漆线",
            "存1线": "存1线",
            "存2线": "存2线",
            "存3线": "存3线",
            "存4线": "存4线",
            "抛丸线": "抛丸线",
        }
        if line_name in mapping:
            return mapping[line_name]
        if line_name == "机走":
            return "机走棚" if position >= 7 else "机走北"
        if line_name == "调梁":
            return "调梁棚" if position >= 7 else "调梁线北"
        if line_name in {"修1", "修2", "修3", "修4"}:
            return f"{line_name}库内" if position >= 5 else f"{line_name}库外"
        if line_name == "存5线":
            return "存5线南" if position >= 22 else "存5线北"
        if line_name == "洗罐":
            return "洗罐站" if position >= 9 else "洗罐线北"
        return line_name

    @staticmethod
    def resolve_target_segment(raw_end_position: str) -> tuple[str, int, int]:
        raw = raw_end_position.strip()
        outer_start = Terminal.repair_outer_start_position
        outer_end = outer_start + Terminal.repair_outer_capacity - 1
        inner_start = Terminal.repair_inner_start_position
        inner_end = inner_start + Terminal.repair_inner_capacity - 1
        mapping = {
            "预修": ("老预修", 1, 14),
            "老预修": ("老预修", 1, 14),
            "预修线": ("老预修", 1, 14),
            "机库线": ("机库线", 1, 5),
            "机库": ("机库线", 1, 5),
            "机北3": ("机走", 1, 6),
            "机走北": ("机走", 1, 6),
            "机棚": ("机走", 7, 14),
            "机走棚": ("机走", 7, 14),
            "机走": ("机走", 1, 14),
            "机走预修": ("机走", 1, 14),
            "调北": ("调梁", 1, 6),
            "调梁线北": ("调梁", 1, 6),
            "调棚": ("调梁", 7, 17),
            "调梁棚": ("调梁", 7, 17),
            "调梁": ("调梁", 1, 17),
            "调梁库": ("调梁", 1, 17),
            "调梁库外": ("调梁", 1, 6),
            "调梁库内": ("调梁", 7, 17),
            "修1库外": ("修1", outer_start, outer_end),
            "修1": ("修1", inner_start, inner_end),
            "修1库内": ("修1", inner_start, inner_end),
            "修2库外": ("修2", outer_start, outer_end),
            "修2": ("修2", inner_start, inner_end),
            "修2库内": ("修2", inner_start, inner_end),
            "修3库外": ("修3", outer_start, outer_end),
            "修3": ("修3", inner_start, inner_end),
            "修3库内": ("修3", inner_start, inner_end),
            "修4库外": ("修4", outer_start, outer_end),
            "修4": ("修4", inner_start, inner_end),
            "修4库内": ("修4", inner_start, inner_end),
            "轮": ("卸轮线", 1, 4),
            "卸轮线": ("卸轮线", 1, 4),
            "漆": ("喷漆", 1, 9),
            "喷漆": ("喷漆", 1, 9),
            "喷漆库": ("喷漆", 1, 9),
            "喷漆库外": ("喷漆", 1, 3),
            "喷漆库内": ("喷漆", 4, 9),
            "油": ("喷漆", 1, 9),
            "油漆": ("喷漆", 1, 9),
            "油漆线": ("喷漆", 1, 9),
            "油漆库": ("喷漆", 1, 9),
            "油漆库外": ("喷漆", 1, 3),
            "油漆库内": ("喷漆", 1, 9),
            "存1": ("存1线", 1, 9),
            "存1线": ("存1线", 1, 9),
            "存2": ("存2线", 1, 20),
            "存2线": ("存2线", 1, 20),
            "存3": ("存3线", 1, 21),
            "存3线": ("存3线", 1, 21),
            "存4": ("存4线", 1, 25),
            "存4线": ("存4线", 1, 25),
            "存4北": ("存4线", 1, 25),
            "存5北": ("存5线", 1, 21),
            "存5线北": ("存5线", 1, 21),
            "存5南": ("存5线", 22, 33),
            "存5线南": ("存5线", 22, 33),
            "存5": ("存5线", 1, 33),
            "存5线": ("存5线", 1, 33),
            "抛": ("抛丸线", 1, 3),
            "抛丸线": ("抛丸线", 1, 3),
            "洗北": ("洗罐", 1, 8),
            "洗罐线北": ("洗罐", 1, 8),
            "洗南": ("洗罐", 9, 15),
            "洗罐站": ("洗罐", 9, 15),
            "洗罐": ("洗罐", 1, 15),
            "洗罐线": ("洗罐", 1, 15),
            "洗罐线外": ("洗罐", 1, 8),
            "洗罐库外": ("洗罐", 1, 8),
            "洗罐线内": ("洗罐", 9, 15),
            "洗罐库内": ("洗罐", 9, 15),
        }
        return mapping.get(raw, (raw, 1, 2**31 - 1))

    @staticmethod
    def ease_key(group: PlacementGroup) -> tuple:
        return (
            group.is_same_line,
            group.blocker_count,
            group.distance,
            group.chunk_index,
            group.start_index,
            group.source_order,
            len(group.cars),
        )

    @staticmethod
    def compare(x: PlacementGroup | None, y: PlacementGroup | None) -> int:
        if x is y:
            return 0
        if x is None:
            return 1
        if y is None:
            return -1
        left = Terminal.ease_key(x)
        right = Terminal.ease_key(y)
        if left < right:
            return -1
        if left > right:
            return 1
        return 0

    @staticmethod
    def get_candidate_tracks(target_type: str, track_line_dict: dict[str, TrackLine]) -> list[TrackLine]:
        return []

    @staticmethod
    def get_exact_capacity_for_target_type(target_type: str) -> int:
        return 0

    @staticmethod
    def group_by_source(groups: list[PlacementGroup]) -> dict[str, list[PlacementGroup]]:
        result: dict[str, list[PlacementGroup]] = {}
        for group in groups:
            result.setdefault(group.source_line.name, []).append(group)
        return result

    GenerateEndSheet = generate_end_sheet
    OutputFile = output_file
    CopySourceFile = copy_source_file
    AddEndPositionToStartSheet = add_end_position_to_start_sheet
    BuildGlobalDict = build_global_dict
    GetTargetLineSegment = get_target_line_segment
    AssignPositionDirect = assign_position_direct
    BuildAssignedPositionDict = build_assigned_position_dict
    TryBuildForceAwareAssignedPositionDict = try_build_force_aware_assigned_position_dict
    BuildGroupPlacementPlans = build_group_placement_plans
    TryAssignUnanchoredGroup = try_assign_unanchored_group
    GetFreeIntervals = get_free_intervals
    FindCompactRemainPositions = find_compact_remain_positions
    BuildPlacementGroups = build_placement_groups
    CreatePlacementGroup = create_placement_group
    SplitPlacementGroup = split_placement_group
    GetDistance = get_distance
    GetSourcePriority = get_source_priority
    NormalizeTrackNameForDistance = normalize_track_name_for_distance
    MapEndLineNameByPosition = map_end_line_name_by_position
    ResolveTargetSegment = resolve_target_segment
    ResolveGroupSegment = resolve_group_segment
    Compare = compare
    GetCandidateTracks = get_candidate_tracks
    GetExactCapacityForTargetType = get_exact_capacity_for_target_type
