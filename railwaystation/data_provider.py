from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from .core import Car, TrackLine, TrackLineName
from .terminal_forced_position_mapper import TerminalForcedPositionMapper


class DataProvider:
    @staticmethod
    def get_row_value(row: tuple, index: int):
        if index < 0 or index >= len(row):
            return None
        return row[index]

    @staticmethod
    def is_effectively_empty_row(row: tuple) -> bool:
        return all(value is None or str(value).strip() == "" for value in row)

    @staticmethod
    def get_sheet(workbook, sheet_name: str):
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
        lowered = sheet_name.lower()
        for name in workbook.sheetnames:
            if name.lower() == lowered:
                return workbook[name]
        raise KeyError(f"Worksheet {sheet_name} does not exist.")

    @staticmethod
    def init_for_terminal(path: str) -> tuple[dict[str, TrackLine], dict[str, Car]]:
        from .terminal import Terminal

        track_lines = {line.value: TrackLine(name=line.value) for line in TrackLineName if line != TrackLineName.train}
        cars: dict[str, Car] = {}
        workbook = load_workbook(path)
        sheet = DataProvider.get_sheet(workbook, "Start_with_end")
        target_line_column = DataProvider.get_column_index(sheet, "末尾位置")
        if target_line_column <= 0:
            target_line_column = sheet.max_column
        force_target_position_column = DataProvider.get_column_index(sheet, "强制对位")
        closed_door_column = DataProvider.get_column_index(sheet, "关门车")
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=0):
            if idx == 0:
                continue
            if DataProvider.is_effectively_empty_row(row):
                continue
            original_track_name = Terminal.normalize_track_name_for_distance(str(DataProvider.get_row_value(row, 0) or "").strip())
            cell_type = str(DataProvider.get_row_value(row, 2) or "")
            if not cell_type:
                continue
            force_raw = DataProvider.get_row_value(row, force_target_position_column - 1) if force_target_position_column > 0 else None
            force_target_position_text = "" if force_raw is None else str(force_raw).strip()
            target_line_name = str(DataProvider.get_row_value(row, target_line_column - 1) or "") if target_line_column > 0 else ""
            possible_target_line_names = DataProvider.parse_possible_target_line_names(target_line_name)
            selected_target_line_name = ""
            target_min_position = 0
            target_max_position = -1
            if len(possible_target_line_names) == 1:
                selected_target_line_name, target_min_position, target_max_position = Terminal.resolve_target_segment(possible_target_line_names[0])
            closed_door_raw = DataProvider.get_row_value(row, closed_door_column - 1) if closed_door_column > 0 else None
            car = Car(
                type=cell_type,
                no=str(DataProvider.get_row_value(row, 3) or ""),
                origin_line_name=original_track_name,
                origin_line_position=int(DataProvider.get_row_value(row, 1) or 0),
                target_line_name=selected_target_line_name,
                possible_target_line_names=possible_target_line_names,
                target_min_position=target_min_position,
                target_max_position=target_max_position,
                force_target_position_text=force_target_position_text,
                is_closed_door=DataProvider.parse_force_flag(closed_door_raw),
                is_force_target_position=False,
                fixed_target_line_position=-1,
                target_line_position=-1,
            )
            if len(possible_target_line_names) == 1:
                TerminalForcedPositionMapper.apply_to_car(car, possible_target_line_names[0])
            cars[car.no] = car
            origin_track_line = track_lines.get(original_track_name)
            if origin_track_line is None:
                raise RuntimeError(f"Start_with_end 中存在无法识别的起始股道：{original_track_name}，车号={car.no}")
            origin_track_line.unshift_current(car)
        return track_lines, cars

    @staticmethod
    def init_track_lines(track_line_capacity: dict[str, Decimal]) -> dict[str, TrackLine]:
        track_lines: dict[str, TrackLine] = {}
        for line_name in TrackLineName:
            if line_name == TrackLineName.train:
                continue
            track_lines[line_name.value] = TrackLine(name=line_name.value, ori_capacity=track_line_capacity[line_name.value])
        return track_lines

    @staticmethod
    def init_cars(track_lines: dict[str, TrackLine], path: str) -> dict[str, Car]:
        cars: dict[str, Car] = {}
        workbook = load_workbook(path)
        sheet_start = DataProvider.get_sheet(workbook, "Start")
        closed_door_column = DataProvider.get_column_index(sheet_start, "关门车")
        for idx, row in enumerate(sheet_start.iter_rows(values_only=True), start=0):
            if idx == 0:
                continue
            if DataProvider.is_effectively_empty_row(row):
                continue
            track_line, line_name, line_name_second = DataProvider.get_car_track_line(
                track_lines, str(DataProvider.get_row_value(row, 0) or "")
            )
            cell_type = str(DataProvider.get_row_value(row, 2) or "")
            if not cell_type:
                continue
            car = Car(
                type=cell_type,
                no=str(DataProvider.get_row_value(row, 3) or ""),
                origin_line_name=line_name,
                origin_line_name_second=line_name_second,
                origin_line_position=int(DataProvider.get_row_value(row, 1) or 0),
                is_heavy=str(DataProvider.get_row_value(row, 4) or "") == "重",
                is_weigh=int(DataProvider.get_row_value(row, 6) or 0) == 1,
                is_closed_door=DataProvider.parse_force_flag(
                    DataProvider.get_row_value(row, closed_door_column - 1) if closed_door_column > 0 else None
                ),
            )
            cars[car.no] = car
            track_line.unshift_current(car)
        sheet_end = DataProvider.get_sheet(workbook, "End_generated")
        for idx, row in enumerate(sheet_end.iter_rows(values_only=True), start=0):
            if idx == 0:
                continue
            if DataProvider.is_effectively_empty_row(row):
                continue
            car_no = str(DataProvider.get_row_value(row, 3) or "")
            if not car_no:
                continue
            track_line, line_name, line_name_second = DataProvider.get_car_track_line(
                track_lines, str(DataProvider.get_row_value(row, 0) or "")
            )
            if car_no in cars:
                car = cars[car_no]
                car.target_line_name = line_name
                car.target_line_name_second = line_name_second
                car.target_line_position = int(DataProvider.get_row_value(row, 1) or 0)
                track_line.unshift_target(car)
                track_line.unshift_origin_target(car)
        return cars

    @staticmethod
    def get_car_track_line(track_lines: dict[str, TrackLine], line_full_name: str) -> tuple[TrackLine, str, str]:
        line_name = line_full_name
        line_name_second = ""
        if line_name.startswith("修"):
            line_name_second = line_name[2:]
            line_name = line_name[:2]
        if len(line_name) >= 2 and line_name.startswith("存5"):
            line_name_second = line_name[3:]
            line_name = line_name[:3]
        if len(line_name) >= 2:
            prefix = line_name[:2]
            if prefix in {"喷漆", "调梁", "洗罐", "机走"}:
                line_name_second = line_name[2:]
                line_name = line_name[:2]
        track_line = track_lines.get(line_name)
        if track_line is None:
            track_line = TrackLine(name=line_name)
            track_lines[line_name] = track_line
        return track_line, line_name, line_name_second

    @staticmethod
    def get_map_info(map_file_path: str) -> tuple[dict[str, dict[str, int]], dict[str, Decimal]]:
        workbook = load_workbook(map_file_path, data_only=True)
        distance_matrix: dict[str, dict[str, int]] = {}
        try:
            sheet = DataProvider.get_sheet(workbook, "DistanceMatrix")
            target_track_ids = [str(sheet.cell(1, col).value or "").strip() for col in range(2, sheet.max_column + 1)]
            for row in range(2, sheet.max_row + 1):
                source_track_id = str(sheet.cell(row, 1).value or "").strip()
                if not source_track_id:
                    continue
                distance_matrix.setdefault(source_track_id, {})
                for col in range(2, sheet.max_column + 1):
                    target_track_id = target_track_ids[col - 2]
                    cell = sheet.cell(row, col).value
                    if cell is not None and cell != "":
                        distance_value = int(round(float(cell)))
                    elif source_track_id == target_track_id:
                        distance_value = 0
                    else:
                        raise RuntimeError("读取DistanceMatrix时存在无效数据")
                    distance_matrix[source_track_id][target_track_id] = distance_value
        except Exception as exc:
            print(f"读取距离矩阵异常: {exc}")
        capacity_sheet = DataProvider.get_sheet(workbook, "Capacity")
        track_line_capacity: dict[str, Decimal] = {}
        for idx, row in enumerate(capacity_sheet.iter_rows(values_only=True), start=0):
            if idx == 0:
                continue
            track_line_capacity[str(row[0] or "")] = Decimal(str(row[1] or 0))
        return distance_matrix, track_line_capacity

    @staticmethod
    def get_column_index(sheet, header_name: str) -> int:
        for col in range(1, sheet.max_column + 1):
            if str(sheet.cell(1, col).value or "").strip() == header_name:
                return col
        return -1

    @staticmethod
    def parse_force_flag(value) -> bool:
        if value is None:
            return False
        normalized = str(value).strip()
        return normalized in {"1", "是", "true", "True", "yes", "Yes", "y", "Y"}

    @staticmethod
    def parse_possible_target_line_names(raw_target_line_name: str) -> list[str]:
        if not raw_target_line_name or not raw_target_line_name.strip():
            return []
        return [
            item
            for item in dict.fromkeys(
                part.strip()
                for part in raw_target_line_name.split(",")
                if part.strip() and part.strip() != "未找到"
            )
        ]

    GetRowValue = get_row_value
    IsEffectivelyEmptyRow = is_effectively_empty_row
    GetSheet = get_sheet
    InitForTerminal = init_for_terminal
    InitTrackLines = init_track_lines
    InitCars = init_cars
    GetCarTrackLine = get_car_track_line
    GetMapInfo = get_map_info
    GetColumnIndex = get_column_index
    ParseForceFlag = parse_force_flag
    ParsePossibleTargetLineNames = parse_possible_target_line_names
