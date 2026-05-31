from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal

from .core import Car, TrackLine
from .data_provider import DataProvider


@dataclass
class TerminalContext:
    distance_matrix: dict[str, dict[str, int]] = field(default_factory=dict)
    track_lines: dict[str, TrackLine] = field(default_factory=dict)
    cars: dict[str, Car] = field(default_factory=dict)
    track_line_capacity: dict[str, Decimal] = field(default_factory=dict)
    start_position_dict: dict[str, int] = field(default_factory=dict)

    @classmethod
    def build_terminal_context(cls, file_path: str) -> "TerminalContext":
        from openpyxl import load_workbook
        from pathlib import Path

        workbook = load_workbook(file_path, data_only=True)
        sheet = DataProvider.get_sheet(workbook, "end")
        current_dir = Path(file_path).parent
        map_path = current_dir / "map.xlsx"
        distance_matrix, track_line_capacity = DataProvider.get_map_info(str(map_path))
        track_line_dict, car_dict = DataProvider.init_for_terminal(file_path)
        for car in car_dict.values():
            car.target_line_position = car.fixed_target_line_position if car.is_force_target_position else -1
        start_position_dict: dict[str, int] = {}
        from .terminal import Terminal

        for row in sheet.iter_rows(min_row=2, values_only=True):
            track_name = str(row[0] or "").strip()
            position = int(row[1] or 0)
            normalized_name = Terminal.normalize_track_name_for_distance(track_name)
            start_position_dict[normalized_name] = max(start_position_dict.get(normalized_name, 0), position)
        return cls(
            distance_matrix=distance_matrix,
            track_lines=track_line_dict,
            cars=car_dict,
            track_line_capacity=track_line_capacity,
            start_position_dict=start_position_dict,
        )

    DistanceMatrix = property(lambda self: self.distance_matrix)
    TrackLineCapacity = property(lambda self: self.track_line_capacity)
    TrackLineDict = property(lambda self: self.track_lines)
    CarDict = property(lambda self: self.cars)
    StartPositonDict = property(lambda self: self.start_position_dict)
    BuildTerminalContext = build_terminal_context
