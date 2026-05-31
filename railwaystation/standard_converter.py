from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment


@dataclass
class TemplateInfo:
    row_by_line_and_position: dict[tuple[str, int], int] = field(default_factory=dict)
    first_position_by_line: dict[str, int] = field(default_factory=dict)
    start_position_by_line_and_remark: dict[tuple[str, str], int] = field(default_factory=dict)


class StandardCaseConverter:
    TEMP_SUFFIX = "_terminal_export_temp_standardized"
    STANDARD_LAYOUT: tuple[tuple[str, int], ...] = (
        ("老预修", 14),
        ("机库线", 5),
        ("机走", 14),
        ("调梁", 17),
        ("喷漆", 9),
        ("洗罐", 15),
        ("修1", 9),
        ("修2", 9),
        ("修3", 9),
        ("修4", 9),
        ("卸轮线", 4),
        ("存1线", 9),
        ("存2线", 20),
        ("存3线", 21),
        ("存4线", 25),
        ("存5线", 33),
        ("抛丸线", 4),
    )

    @staticmethod
    def convert_case(source_file_path: str, output_file_path: str | None = None, template_path: str | None = None) -> Path:
        source_path = Path(source_file_path).resolve()
        output_path = (
            Path(output_file_path).resolve()
            if output_file_path
            else source_path.with_name(f"{source_path.stem}{StandardCaseConverter.TEMP_SUFFIX}{source_path.suffix}")
        )

        source_workbook = load_workbook(source_path)

        if "Start" not in source_workbook.sheetnames:
            raise RuntimeError(f"源文件不存在 Start 表：{source_path}")
        if "End" not in source_workbook.sheetnames:
            raise RuntimeError(f"源文件不存在 End 表：{source_path}")

        output_workbook = Workbook()
        output_workbook.remove(output_workbook.active)
        StandardCaseConverter.build_standard_sheet(output_workbook.create_sheet("Start"))
        StandardCaseConverter.build_standard_sheet(output_workbook.create_sheet("End"))

        output_start_sheet = output_workbook["Start"]
        output_end_sheet = output_workbook["End"]

        StandardCaseConverter.clear_template_car_data(output_start_sheet)
        StandardCaseConverter.clear_template_car_data(output_end_sheet)

        StandardCaseConverter.convert_sheet(source_workbook["Start"], output_start_sheet)
        StandardCaseConverter.convert_sheet(source_workbook["End"], output_end_sheet)

        StandardCaseConverter.format_sheet(output_start_sheet)
        StandardCaseConverter.format_sheet(output_end_sheet)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        if output_path.exists():
            output_path.unlink()
        output_workbook.save(output_path)

        source_workbook.close()
        output_workbook.close()
        return output_path

    @staticmethod
    def find_template_file(source_path: Path) -> Path:
        candidates: list[Path] = []
        search_roots = [
            source_path.parent,
            Path(__file__).resolve().parent.parent / "Data",
        ]
        seen: set[Path] = set()
        for root in search_roots:
            if root in seen or not root.exists():
                continue
            seen.add(root)
            for path in root.glob("*.xlsx"):
                name = path.name
                if name.startswith("~$"):
                    continue
                if "标准化" in name and "模板" in name:
                    candidates.append(path.resolve())
        if not candidates:
            raise FileNotFoundError(f"未找到标准化模板文件，源文件={source_path}")
        return candidates[0]

    @staticmethod
    def build_standard_sheet(sheet) -> None:
        headers = ("股道", "序号", "车型", "车号", "修程", "备注", "扣车日期", "目标股道", "是否对位", "车辆属性")
        for col, header in enumerate(headers, start=1):
            sheet.cell(1, col).value = header
        row_number = 2
        for line_name, capacity in StandardCaseConverter.STANDARD_LAYOUT:
            for position in range(1, capacity + 1):
                sheet.cell(row_number, 1).value = line_name
                sheet.cell(row_number, 2).value = position
                row_number += 1

    @staticmethod
    def convert_sheet(source_sheet, target_sheet) -> None:
        source_header_row = StandardCaseConverter.find_header_row(source_sheet)
        target_header_row = StandardCaseConverter.find_header_row(target_sheet)

        source_columns = StandardCaseConverter.build_header_column_map(source_sheet, source_header_row)
        target_columns = StandardCaseConverter.build_header_column_map(target_sheet, target_header_row)

        source_line_col = StandardCaseConverter.get_required_column(source_columns, "股道", source_sheet.title)
        source_position_col = StandardCaseConverter.get_required_column(source_columns, "序号", source_sheet.title)
        source_car_no_col = StandardCaseConverter.get_required_column(source_columns, "车号", source_sheet.title)

        target_line_col = StandardCaseConverter.get_required_column(target_columns, "股道", target_sheet.title)
        target_position_col = StandardCaseConverter.get_required_column(target_columns, "序号", target_sheet.title)

        template_info = StandardCaseConverter.build_template_info(
            target_sheet,
            target_header_row,
            target_line_col,
            target_position_col,
        )
        source_line_min_position = StandardCaseConverter.build_source_line_min_position(
            source_sheet,
            source_header_row,
            source_line_col,
            source_position_col,
        )
        source_line_first_occupied_position = StandardCaseConverter.build_source_line_first_occupied_position(
            source_sheet,
            source_header_row,
            source_line_col,
            source_position_col,
            source_car_no_col,
        )

        for row_number in range(source_header_row + 1, source_sheet.max_row + 1):
            old_line_name = str(source_sheet.cell(row_number, source_line_col).value or "").strip()
            car_no = str(source_sheet.cell(row_number, source_car_no_col).value or "").strip()
            if not old_line_name or not car_no:
                continue
            old_position = StandardCaseConverter.try_get_int(source_sheet.cell(row_number, source_position_col).value)
            if old_position is None:
                continue

            new_line_name, new_position = StandardCaseConverter.map_old_position_to_new_position(
                old_line_name,
                old_position,
                source_line_min_position,
                source_line_first_occupied_position,
                template_info,
            )
            target_row_number = template_info.row_by_line_and_position.get((new_line_name, new_position))
            if target_row_number is None:
                raise RuntimeError(
                    f"未找到模板台位：表={source_sheet.title}，旧股道={old_line_name}，旧序号={old_position}，"
                    f"新股道={new_line_name}，新序号={new_position}，车号={car_no}"
                )
            StandardCaseConverter.copy_car_data(
                source_sheet,
                row_number,
                target_sheet,
                target_row_number,
                source_columns,
                target_columns,
            )

    @staticmethod
    def map_old_position_to_new_position(
        old_line_name: str,
        old_position: int,
        source_line_min_position: dict[str, int],
        source_line_first_occupied_position: dict[str, int],
        template_info: TemplateInfo,
    ) -> tuple[str, int]:
        name = old_line_name.strip()
        mapping = {
            "老预修": lambda: StandardCaseConverter.by_line_start("老预修", old_line_name, old_position, source_line_min_position, template_info),
            "预修": lambda: StandardCaseConverter.by_line_start("老预修", old_line_name, old_position, source_line_min_position, template_info),
            "机库线": lambda: StandardCaseConverter.by_line_start("机库线", old_line_name, old_position, source_line_min_position, template_info),
            "机库": lambda: StandardCaseConverter.by_line_start("机库线", old_line_name, old_position, source_line_min_position, template_info),
            "机走预修": lambda: StandardCaseConverter.by_line_start("机走", old_line_name, old_position, source_line_min_position, template_info),
            "机北3": lambda: StandardCaseConverter.by_fixed_start("机走", 1, old_line_name, old_position, source_line_min_position),
            "机走库外": lambda: StandardCaseConverter.by_fixed_start("机走", 1, old_line_name, old_position, source_line_min_position),
            "机棚": lambda: StandardCaseConverter.by_fixed_start("机走", 7, old_line_name, old_position, source_line_min_position),
            "机走库内": lambda: StandardCaseConverter.by_fixed_start("机走", 7, old_line_name, old_position, source_line_min_position),
            "调梁库": lambda: StandardCaseConverter.by_line_start("调梁", old_line_name, old_position, source_line_min_position, template_info),
            "调梁库外": lambda: StandardCaseConverter.by_fixed_old_start_and_new_start("调梁", 1, 1, old_position),
            "调北": lambda: StandardCaseConverter.by_fixed_old_start_and_new_start("调梁", 1, 1, old_position),
            "调梁库内": lambda: StandardCaseConverter.by_fixed_old_start_and_new_start("调梁", 6, 7, old_position),
            "调棚": lambda: StandardCaseConverter.by_fixed_old_start_and_new_start("调梁", 6, 7, old_position),
            "喷漆库": lambda: StandardCaseConverter.by_line_start("喷漆", old_line_name, old_position, source_line_min_position, template_info),
            "喷漆库外": lambda: StandardCaseConverter.by_line_start("喷漆", old_line_name, old_position, source_line_min_position, template_info),
            "喷漆库内": lambda: StandardCaseConverter.by_line_start("喷漆", old_line_name, old_position, source_line_min_position, template_info),
            "修1库外": lambda: StandardCaseConverter.by_fixed_start("修1", 1, old_line_name, old_position, source_line_first_occupied_position),
            "修1库内": lambda: StandardCaseConverter.by_fixed_start("修1", 5, old_line_name, old_position, source_line_min_position),
            "修2库外": lambda: StandardCaseConverter.by_fixed_start("修2", 1, old_line_name, old_position, source_line_first_occupied_position),
            "修2库内": lambda: StandardCaseConverter.by_fixed_start("修2", 5, old_line_name, old_position, source_line_min_position),
            "修3库外": lambda: StandardCaseConverter.by_fixed_start("修3", 1, old_line_name, old_position, source_line_first_occupied_position),
            "修3库内": lambda: StandardCaseConverter.by_fixed_start("修3", 5, old_line_name, old_position, source_line_min_position),
            "修4库外": lambda: StandardCaseConverter.by_fixed_start("修4", 1, old_line_name, old_position, source_line_first_occupied_position),
            "修4库内": lambda: StandardCaseConverter.by_fixed_start("修4", 5, old_line_name, old_position, source_line_min_position),
            "卸轮线": lambda: StandardCaseConverter.by_line_start("卸轮线", old_line_name, old_position, source_line_min_position, template_info),
            "轮": lambda: StandardCaseConverter.by_line_start("卸轮线", old_line_name, old_position, source_line_min_position, template_info),
            "存1": lambda: StandardCaseConverter.by_line_start("存1线", old_line_name, old_position, source_line_min_position, template_info),
            "存2": lambda: StandardCaseConverter.by_line_start("存2线", old_line_name, old_position, source_line_min_position, template_info),
            "存3": lambda: StandardCaseConverter.by_line_start("存3线", old_line_name, old_position, source_line_min_position, template_info),
            "存4": lambda: StandardCaseConverter.by_line_start("存4线", old_line_name, old_position, source_line_min_position, template_info),
            "存1线": lambda: StandardCaseConverter.by_line_start("存1线", old_line_name, old_position, source_line_min_position, template_info),
            "存2线": lambda: StandardCaseConverter.by_line_start("存2线", old_line_name, old_position, source_line_min_position, template_info),
            "存3线": lambda: StandardCaseConverter.by_line_start("存3线", old_line_name, old_position, source_line_min_position, template_info),
            "存4线": lambda: StandardCaseConverter.by_line_start("存4线", old_line_name, old_position, source_line_min_position, template_info),
            "存4北": lambda: StandardCaseConverter.by_line_start("存4线", old_line_name, old_position, source_line_min_position, template_info),
            "存5线": lambda: StandardCaseConverter.by_line_start("存5线", old_line_name, old_position, source_line_min_position, template_info),
            "存5线北": lambda: StandardCaseConverter.by_fixed_start("存5线", 1, old_line_name, old_position, source_line_min_position),
            "存5北": lambda: StandardCaseConverter.by_fixed_start("存5线", 1, old_line_name, old_position, source_line_min_position),
            "存5线南": lambda: StandardCaseConverter.by_fixed_start("存5线", 22, old_line_name, old_position, source_line_min_position),
            "存5南": lambda: StandardCaseConverter.by_fixed_start("存5线", 1, old_line_name, old_position, source_line_min_position),
            "抛丸线": lambda: StandardCaseConverter.by_line_start("抛丸线", old_line_name, old_position, source_line_min_position, template_info),
            "抛": lambda: StandardCaseConverter.by_line_start("抛丸线", old_line_name, old_position, source_line_min_position, template_info),
            "洗罐线": lambda: StandardCaseConverter.by_line_start("洗罐", old_line_name, old_position, source_line_min_position, template_info),
            "洗罐库外": lambda: StandardCaseConverter.by_fixed_start("洗罐", 1, old_line_name, old_position, source_line_min_position),
            "洗罐线外": lambda: StandardCaseConverter.by_fixed_start("洗罐", 1, old_line_name, old_position, source_line_min_position),
            "洗罐库内": lambda: StandardCaseConverter.by_fixed_start("洗罐", 9, old_line_name, old_position, source_line_min_position),
            "洗罐线内": lambda: StandardCaseConverter.by_fixed_start("洗罐", 9, old_line_name, old_position, source_line_min_position),
            "洗南": lambda: StandardCaseConverter.by_fixed_start("洗罐", 9, old_line_name, old_position, source_line_min_position),
            "洗北": lambda: StandardCaseConverter.by_fixed_start("洗罐", 1, old_line_name, old_position, source_line_min_position),
            "机走": lambda: StandardCaseConverter.by_line_start("机走", old_line_name, old_position, source_line_min_position, template_info),
            "调梁": lambda: StandardCaseConverter.by_line_start("调梁", old_line_name, old_position, source_line_min_position, template_info),
            "修1": lambda: StandardCaseConverter.direct("修1", old_position),
            "修2": lambda: StandardCaseConverter.direct("修2", old_position),
            "修3": lambda: StandardCaseConverter.direct("修3", old_position),
            "修4": lambda: StandardCaseConverter.direct("修4", old_position),
            "喷漆": lambda: StandardCaseConverter.by_line_start("喷漆", old_line_name, old_position, source_line_min_position, template_info),
            "油": lambda: StandardCaseConverter.by_line_start("喷漆", old_line_name, old_position, source_line_min_position, template_info),
            "洗罐": lambda: StandardCaseConverter.by_line_start("洗罐", old_line_name, old_position, source_line_min_position, template_info),
            "存5": lambda: StandardCaseConverter.by_line_start("存5线", old_line_name, old_position, source_line_min_position, template_info),
        }
        resolver = mapping.get(name)
        if resolver is not None:
            return resolver()
        return StandardCaseConverter.by_line_start(name, old_line_name, old_position, source_line_min_position, template_info)

    @staticmethod
    def by_fixed_old_start_and_new_start(new_line_name: str, old_start_position: int, new_start_position: int, old_position: int) -> tuple[str, int]:
        relative = old_position - old_start_position + 1
        return new_line_name, new_start_position + relative - 1

    @staticmethod
    def direct(new_line_name: str, old_position: int) -> tuple[str, int]:
        return new_line_name, old_position

    @staticmethod
    def by_fixed_start(
        new_line_name: str,
        new_start_position: int,
        old_line_name: str,
        old_position: int,
        source_line_min_position: dict[str, int],
    ) -> tuple[str, int]:
        relative = StandardCaseConverter.get_relative_position(old_line_name, old_position, source_line_min_position)
        return new_line_name, new_start_position + relative - 1

    @staticmethod
    def by_line_start(
        new_line_name: str,
        old_line_name: str,
        old_position: int,
        source_line_min_position: dict[str, int],
        template_info: TemplateInfo,
    ) -> tuple[str, int]:
        start_position = template_info.first_position_by_line.get(new_line_name, 1)
        return StandardCaseConverter.by_fixed_start(
            new_line_name,
            start_position,
            old_line_name,
            old_position,
            source_line_min_position,
        )

    @staticmethod
    def get_relative_position(old_line_name: str, old_position: int, source_line_min_position: dict[str, int]) -> int:
        min_position = source_line_min_position.get(old_line_name, old_position)
        return old_position - min_position + 1

    @staticmethod
    def copy_car_data(source_sheet, source_row: int, target_sheet, target_row: int, source_columns: dict[str, int], target_columns: dict[str, int]) -> None:
        for header_name in ("车型", "车号", "修程", "扣车日期", "目标股道", "是否对位", "车辆属性"):
            source_col = source_columns.get(header_name)
            target_col = target_columns.get(header_name)
            if source_col is None or target_col is None:
                continue
            target_sheet.cell(target_row, target_col).value = source_sheet.cell(source_row, source_col).value

    @staticmethod
    def build_template_info(sheet, header_row: int, line_column: int, position_column: int) -> TemplateInfo:
        info = TemplateInfo()
        segment_remark_columns = StandardCaseConverter.find_all_header_columns(sheet, header_row, "备注")
        segment_remark_column = segment_remark_columns[-1] if segment_remark_columns else 0
        for row_number in range(header_row + 1, sheet.max_row + 1):
            line_name = str(sheet.cell(row_number, line_column).value or "").strip()
            if not line_name:
                continue
            position = StandardCaseConverter.try_get_int(sheet.cell(row_number, position_column).value)
            if position is None:
                continue
            info.row_by_line_and_position[(line_name, position)] = row_number
            first_position = info.first_position_by_line.get(line_name)
            if first_position is None or position < first_position:
                info.first_position_by_line[line_name] = position
            if segment_remark_column > 0:
                remark = str(sheet.cell(row_number, segment_remark_column).value or "").strip()
                if remark:
                    info.start_position_by_line_and_remark[(line_name, remark)] = position
        return info

    @staticmethod
    def build_source_line_min_position(sheet, header_row: int, line_column: int, position_column: int) -> dict[str, int]:
        result: dict[str, int] = {}
        for row_number in range(header_row + 1, sheet.max_row + 1):
            line_name = str(sheet.cell(row_number, line_column).value or "").strip()
            if not line_name:
                continue
            position = StandardCaseConverter.try_get_int(sheet.cell(row_number, position_column).value)
            if position is None:
                continue
            current = result.get(line_name)
            if current is None or position < current:
                result[line_name] = position
        return result

    @staticmethod
    def build_source_line_first_occupied_position(
        sheet,
        header_row: int,
        line_column: int,
        position_column: int,
        car_no_column: int,
    ) -> dict[str, int]:
        result: dict[str, int] = {}
        for row_number in range(header_row + 1, sheet.max_row + 1):
            line_name = str(sheet.cell(row_number, line_column).value or "").strip()
            car_no = str(sheet.cell(row_number, car_no_column).value or "").strip()
            if not line_name or not car_no:
                continue
            position = StandardCaseConverter.try_get_int(sheet.cell(row_number, position_column).value)
            if position is None:
                continue
            current = result.get(line_name)
            if current is None or position < current:
                result[line_name] = position
        return result

    @staticmethod
    def clear_template_car_data(sheet) -> None:
        header_row = StandardCaseConverter.find_header_row(sheet)
        columns = StandardCaseConverter.build_header_column_map(sheet, header_row)
        clear_headers = ("车型", "车号", "修程", "扣车日期", "目标股道", "是否对位", "车辆属性")
        for header in clear_headers:
            col = columns.get(header)
            if col is None:
                continue
            for row_number in range(header_row + 1, sheet.max_row + 1):
                sheet.cell(row_number, col).value = None
        for remark_col in StandardCaseConverter.find_all_header_columns(sheet, header_row, "备注"):
            for row_number in range(header_row + 1, sheet.max_row + 1):
                sheet.cell(row_number, remark_col).value = None

    @staticmethod
    def find_all_header_columns(sheet, header_row: int, header_name: str) -> list[int]:
        result: list[int] = []
        for col in range(1, sheet.max_column + 1):
            if str(sheet.cell(header_row, col).value or "").strip() == header_name:
                result.append(col)
        return result

    @staticmethod
    def format_sheet(sheet) -> None:
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[sheet.cell(1, col).column_letter].width = 12
        for row in range(1, sheet.max_row + 1):
            sheet.row_dimensions[row].height = 20
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def find_header_row(sheet) -> int:
        for row_number in range(1, sheet.max_row + 1):
            values = [
                str(sheet.cell(row_number, col).value or "").strip()
                for col in range(1, sheet.max_column + 1)
                if str(sheet.cell(row_number, col).value or "").strip()
            ]
            if {"股道", "序号", "车型", "车号"}.issubset(set(values)):
                return row_number
        raise RuntimeError(f"工作表 {sheet.title} 中未找到表头行")

    @staticmethod
    def build_header_column_map(sheet, header_row: int) -> dict[str, int]:
        result: dict[str, int] = {}
        for col in range(1, sheet.max_column + 1):
            header = str(sheet.cell(header_row, col).value or "").strip()
            if header and header not in result:
                result[header] = col
        return result

    @staticmethod
    def get_required_column(columns: dict[str, int], header_name: str, sheet_name: str) -> int:
        col = columns.get(header_name)
        if col is None:
            raise RuntimeError(f"工作表 {sheet_name} 缺少必要列：{header_name}")
        return col

    @staticmethod
    def try_get_int(value) -> int | None:
        if value is None:
            return None
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        text = str(value).strip()
        if not text:
            return None
        try:
            return int(float(text))
        except ValueError:
            return None
