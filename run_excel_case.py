from __future__ import annotations

import argparse
import copy
import json
from pathlib import Path

from railwaystation import BackwardConstructionAlgorithm, DataProvider, StandardCaseConverter, Terminal


def apply_csharp_export_postprocess(operations):
    normalized = copy.deepcopy(operations)
    for operation in normalized:
        operation.train_cars.reverse()
        if operation.action.value == "Put":
            operation.move_cars.reverse()
    return normalized


def _serialize_decimal(value):
    try:
        integral = value.to_integral_value()
        if value == integral:
            return int(integral)
    except AttributeError:
        pass
    return float(value)


def car_to_payload(car, include_debug_fields=False):
    payload = {
        "No": car.no,
        "Type": car.type,
        "Length": _serialize_decimal(car.length),
        "CurrentLineId": car.current_line_id,
        "CurrentLineName": car.current_line_name,
        "OriginLineName": car.origin_line_name,
        "OriginLineName_Second": car.origin_line_name_second,
        "OriginLinePosition": car.origin_line_position,
        "PossibleTargetLineNames": list(car.possible_target_line_names),
        "TargetMinPosition": car.target_min_position,
        "TargetMaxPosition": car.target_max_position,
        "TargetLineId": car.target_line_id,
        "TargetLineName": car.target_line_name,
        "TargetLineName_Second": car.target_line_name_second,
        "TargetLinePosition": car.target_line_position,
        "IsForceTargetPosition": car.is_force_target_position,
        "FixedTargetLinePosition": car.fixed_target_line_position,
        "ForceTargetPositionText": car.force_target_position_text,
        "AllowedTargetLinePositions": list(car.allowed_target_line_positions),
        "IsClosedDoor": car.is_closed_door,
        "IsHeavy": car.is_heavy,
        "IsWeigh": car.is_weigh,
        "RemainOriginTargetCars": [],
    }
    if include_debug_fields:
        payload["CurrentDepth"] = car.current_depth
        payload["IsWeighed"] = car.is_weighed
    payload["RemainOriginTargetCars"] = [
        car_to_payload(item, include_debug_fields=include_debug_fields)
        for item in car.remain_origin_target_cars
    ]
    return payload


def build_result_payload(operations, include_debug_fields=False):
    return [
        {
            "Index": op.index,
            "LineName": op.line_name,
            "Action": op.action.value,
            "MoveCarCount": op.move_car_count,
            "TrainCarsCount": op.train_cars_count,
            "LineCarsBeforCount": op.line_cars_befor_count,
            "LineCarsAfterCount": op.line_cars_after_count,
            "MoveCars": [car_to_payload(car, include_debug_fields=include_debug_fields) for car in op.move_cars],
            "TrainCars": [car_to_payload(car, include_debug_fields=include_debug_fields) for car in op.train_cars],
            "LineCarsBefore": [car_to_payload(car, include_debug_fields=include_debug_fields) for car in op.line_cars_before],
            "LineCarsAfter": [car_to_payload(car, include_debug_fields=include_debug_fields) for car in op.line_cars_after],
        }
        for op in operations
    ]


def resolve_working_case_path(case_path: Path, prepare_terminal: bool) -> Path:
    from openpyxl import load_workbook

    if prepare_terminal:
        Terminal.generate_end_sheet(str(case_path))
        return StandardCaseConverter.convert_case(str(case_path))

    workbook = load_workbook(case_path, read_only=True)
    try:
        has_end_generated = "End_generated" in workbook.sheetnames
    finally:
        workbook.close()

    if has_end_generated:
        return case_path

    raise RuntimeError(
        "输入文件缺少 End_generated sheet。"
        "当前默认严格使用外部现成终点以对齐 C# direct_solver_only；"
        "如果你要先生成终点再求解，请显式传入 --prepare-terminal。"
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Run a RailwayStation_V2_py Excel case.")
    parser.add_argument("--file", required=True, type=Path, help="Case xlsx path.")
    parser.add_argument("--map", required=True, type=Path, help="map.xlsx path.")
    parser.add_argument("--output", required=True, type=Path, help="Output json path.")
    parser.add_argument(
        "--prepare-terminal",
        action="store_true",
        help="Generate Start_with_end and End_generated before solving. Disabled by default so the solver uses the workbook's existing End_generated and stays aligned with C# direct_solver_only.",
    )
    parser.add_argument(
        "--apply-csharp-export-postprocess",
        action=argparse.BooleanOptionalAction,
        default=False,
        help="Apply legacy export-side reverse logic. Disabled by default because RailwayStation_V2 parity should preserve the raw operation order; use --apply-csharp-export-postprocess only when you explicitly need the older reversed export format.",
    )
    parser.add_argument(
        "--include-debug-fields",
        action=argparse.BooleanOptionalAction,
        default=False,
        help="Include non-C#-serialized debug fields such as CurrentDepth and IsWeighed.",
    )
    args = parser.parse_args()

    case_path = args.file.resolve()
    map_path = args.map.resolve()
    output_path = args.output.resolve()

    working_case_path = resolve_working_case_path(case_path, args.prepare_terminal)

    distance_matrix, track_line_capacity = DataProvider.get_map_info(str(map_path))
    track_lines = DataProvider.init_track_lines(track_line_capacity)
    cars = DataProvider.init_cars(track_lines, str(working_case_path))
    operations = BackwardConstructionAlgorithm(track_lines, cars, distance_matrix).run()
    if args.apply_csharp_export_postprocess:
        operations = apply_csharp_export_postprocess(operations)

    output_path.write_text(
        json.dumps(build_result_payload(operations, include_debug_fields=args.include_debug_fields), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"operations={len(operations)}")
    print(f"case_used={working_case_path}")
    print(f"output={output_path}")


if __name__ == "__main__":
    main()
